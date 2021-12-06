import xlrd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
import numpy as np
import copy
import matplotlib.pyplot as plt
import networkx as nx
from networkx.algorithms.distance_measures import diameter
dicnode={}
dicway={}
diccity={}
# Create a dictionary of indicator names and serial numbers
name={}
name[1]="betweenness"
name[2]="flow betweenness"
name[3]="sensitivity"
ty=['o','v','x']

# Create a dictionary of city names and index numbers
wbc=load_workbook('Tables for storing dictionaries')
sheet_names = wbc.sheetnames
table2=wbc[sheet_names[0]]
for r1 in range(2,table2.max_row+1):   
    dicnode[table2.cell(row=r1,column=1).value]=[]
    diccity[table2.cell(row=r1,column=2).value-1]=table2.cell(row=r1,column=1).value

yea=2018/2017
# Read the adjacency table and create two hash table dictionaries,
#  a node-line table and a line-connected edge table
filename='data in folders'
wbt=load_workbook(filename)
sheet_names = wbt.sheetnames
table1=wbt[sheet_names[0]]

for r2 in range(1,table1.max_row+1):
    # Set of nodes through which trains pass
    str1=table1.cell(row=r2,column=2).value 
    # train numbers
    num=table1.cell(row=r2,column=1).value 
    ss=str1.split()
    if len(ss)<=1:
        continue
    dicway[num]=[]
    if ss[0] in dicnode:
         # Add this train number to the list of attributes of node 0
        dicnode[ss[0]].append(num)
    for cc in range(1,len(ss)):
        if ss[cc] in dicnode:
            dicnode[ss[cc]].append(num)
            if ss[cc]==ss[cc-1] :
                continue   
            edge=(ss[cc],ss[cc-1])   
            # Add an edge to the list of attributes for this train
            dicway[num].append(edge) 

# Read the current year's adjacency matrix to build the network
filename2='AccessData'
wb= xlrd.open_workbook(filename2)
table=wb.sheets()[0] 
n=table.nrows-1   
# Add the city name to the node name
point = []
j=1
while j<=n:
    point.append(table.cell(0,j).value)
    j=j+1

# Putting excel data into a matrix

matrix = np.zeros((n, n)) 
for h in range(n):
    for l in range(n):
        matrix[h, l] = table.cell(h+1,l+1).value        

G=nx.Graph()
for p in range(n):
    G.add_node(point[p])
for r3 in range(n):
    for c in range(r3+1,n):
        if matrix[r3][c] > 0:
            G.add_edge(diccity[r3],diccity[c],weight=matrix[r3][c])

#Find the initial giant component size
Gc = sorted(nx.connected_components(G), key=len, reverse=True)       
GCC=G.subgraph(Gc[0])#MaxConnectionComponent
size0 =GCC.number_of_nodes()
E0=nx.global_efficiency(G)


# Remove nodes from the network, 
# calculate the maximum connectivity component size of the new network, 
# and record it in a list/array
result=[]
plt.figure(dpi=400,figsize=(4.4,4))
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

for k in range(1,4):
    dicwayn=dicway.copy()
    re=0
    result.clear()
    Gnow=copy.deepcopy(G)
    # Read the node evaluation ranking and determine the order of deletion of nodes
    cen=wb.sheets()[k]

    for i in range(0,cen.nrows):
        Gc.clear()
        nownode=cen.cell(i,0).value
        # Find the line where the node is located
        waylist=dicnode[nownode] 
        numin=0
        for way in waylist:
            if way in dicwayn:
                # Find the edge formed by the line
                edgelist=dicwayn[way] 
                for edge1 in edgelist :
                    try:
                        we=Gnow[edge1[0]][edge1[1]]['weight']                       
                        if we==1:Gnow.remove_edge(edge1[0],edge1[1])
                        #delete corresponding edge
                        else:Gnow[edge1[0]][edge1[1]]['weight'] -= 1 
                    except KeyError:
                        print(Gnow.edges.data())
                        print(nownode,way,edge1,edgelist)
                        exit()    
                    numin += 1
                del dicwayn[way]                       
        Gnow.remove_node(nownode)
        
        Gc = sorted(nx.connected_components(Gnow), key=len, reverse=True)     
        # Maximum connected component after removal of nodes
        GCC=Gnow.subgraph(Gc[0])
        
        siGCC=GCC.number_of_nodes()
        if siGCC==1:break        
        Et=nx.global_efficiency(Gnow)
        #The note is network efficiency
        # re += Et/E0
        # result.append(Et/E0)
        result.append(siGCC/size0)
        re += siGCC/size0
        print(siGCC/size0)
    print(re)
    x=range(1,len(result)+1)
    plt.plot(x,result,ty[k-1],label=name[k],markersize=4)

plt.legend(fontsize=7)
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False
titles=str(yea)
plt.xlabel("n",fontsize=8)
plt.ylabel("relative size of giant component",fontsize=8)
# plt.ylabel("efficiency",fontsize=8)
# plt.title(titles,fontsize=11)
x_ticks = np.arange(0, 16, 1)
y_ticks = np.arange(0, 1.1, 0.1)
plt.xticks(x_ticks,fontsize=7)
plt.yticks(y_ticks,fontsize=7)
plt.show()
print("finished!")












